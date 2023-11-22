from openpyxl import Workbook

wb = Workbook()
#ws = wb.active
ws = wb.create_sheet()
ws.title = "MySheet" 
ws.sheet_properties.tabColor = "ff66ff" #탭 색상 변경

ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet",2)

print(wb.sheetnames)

#wb.save('sample.xlsx')
#wb.close()